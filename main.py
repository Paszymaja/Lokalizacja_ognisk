import openpyxl
import easygui
import numpy as np

path = easygui.fileopenbox()
wb = openpyxl.load_workbook(path)
arkusz = wb['Arkusz1']
wave_speed = arkusz['A14'].value
nr_w = 1
nr_w = int(input('Podaj numer wstrzasu: '))


def wczytanie_danych(sheet, numer_wstrzasu):
    tab = []
    tab2 = []
    dict_nr = {
        1: 7,
        2: 8,
        3: 9,
        4: 10
    }
    for row_i in range(2, 10):
        for column_i in range(2, 5):
            x1 = float(sheet.cell(row_i, column_i).value)
            tab.append(x1)
        x_w = float(sheet.cell(row_i, dict_nr.get(numer_wstrzasu)).value)
        tab2.append(x_w)
    tab2 = np.array(tab2).reshape(8, 1)
    mat = np.array(tab).reshape(8, 3)
    mat = np.append(mat, tab2, axis=1)
    return mat


def linearyzjaca(macierz, v):
    tab = []
    for i in range(0, 8):
        for j in range(0, 4):
            if j != 3:
                tab.append((-2*(macierz[i][j])+2*macierz[0][j]))
            else:
                tab.append(pow(v, 2)*(2*macierz[i][j]-2*macierz[0][j]))
    mat = np.array(tab).reshape(8, 4)
    mat = np.delete(mat, 0, 0)
    return mat


def macierz_b(macierz, v):
    tab = []
    for i in range(0, 8):
        a = 0
        b = 0
        for j in range(0, 4):
            if j != 3:
                a = a+(-(pow(macierz[i][j], 2)) + pow(macierz[0][j], 2))
            else:
                b = b+(pow(v, 2)*(pow(macierz[i][j], 2) - pow(macierz[0][j], 2)))
        tab.append(a+b)
    mat = np.array(tab).reshape(8, 1)
    mat = np.delete(mat, 0, 0)
    return mat


def zapis(sheet, table):
    i = 0
    for j in range(14, 18):
        sheet.cell(row=j, column=4).value = table[i][0]
        print(wynik[i][0])
        i=i+1
    wb.save('wb.xlsx')


macierzA = wczytanie_danych(arkusz, nr_w)


macierzB = macierz_b(macierzA, wave_speed)
macierzA = linearyzjaca(macierzA, wave_speed)


macierzT = macierzA.transpose()


macierzA = np.dot(macierzT, macierzA)
macierzB = np.dot(macierzT, macierzB)
print(macierzA, "\n")
print(macierzB, "\n")

wynik = np.linalg.solve(macierzA, macierzB)
zapis(arkusz, wynik)
















